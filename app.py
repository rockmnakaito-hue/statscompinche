import streamlit as st
import pandas as pd
import unicodedata
from io import BytesIO
from openpyxl import load_workbook
import os
import datetime
import shutil

# =====================
# Función de normalización
# =====================
def normalizar(nombre):
    if pd.isna(nombre):
        return ""
    nombre = str(nombre).strip().lower()
    nombre = ''.join(
        c for c in unicodedata.normalize('NFD', nombre)
        if unicodedata.category(c) != 'Mn'
    )
    return nombre

# =====================
# Rutas de archivos base
# =====================
HORARIO_PATH = "data/horario.xlsx"
ARCHIVO_BASE_PATH = "data/archivo_a.xlsx"
BACKUP_FOLDER = "data/backups"

# Crear carpeta de backups si no existe
os.makedirs(BACKUP_FOLDER, exist_ok=True)

# =====================
# Cargar horario
# =====================
@st.cache_data
def cargar_horario():
    if not os.path.exists(HORARIO_PATH):
        st.error(f"No se encontró el archivo de horario en: {HORARIO_PATH}")
        st.stop()
    df = pd.read_excel(HORARIO_PATH, sheet_name="Turnos")
    df["Nombres"] = df["Nombres"].astype(str)
    return df

# =====================
# Función principal para generar archivo
# =====================
def generar_excel(horario, df_csv, dia, turno):
    fila = horario[(horario["Día"] == dia) & (horario["Turno"] == turno)]
    if fila.empty:
        st.error("No se encontró el día o turno en el horario.")
        return None

    agentes_turno = [a.strip() for a in fila.iloc[0]["Nombres"].split(",") if a.strip()]
    agentes_norm = [normalizar(a) for a in agentes_turno]

    if "First Name" not in df_csv.columns:
        st.error("El CSV debe tener la columna 'First Name'")
        return None

    # Normalizar solo First Name
    df_csv["Nombre_norm"] = df_csv["First Name"].apply(normalizar)

    # Filtrar solo agentes del turno
    datos = df_csv[df_csv["Nombre_norm"].isin(agentes_norm)]

    detectados = datos["First Name"].unique().tolist()
    no_detectados = [a for a in agentes_turno if normalizar(a) not in datos["Nombre_norm"].values]

    # Mostrar tabla de agentes
    preview_df = pd.DataFrame({
        "Agente": agentes_turno,
        "Detectado en CSV": ["✅" if normalizar(a) in df_csv["Nombre_norm"].values else "❌" for a in agentes_turno]
    })

    def color_detectado(val):
        color = 'green' if val == '✅' else 'red'
        return f'color: {color}; font-weight: bold'

    st.subheader("🕒 Estado de agentes en turno")
    st.dataframe(preview_df.style.applymap(color_detectado, subset=["Detectado en CSV"]))

    # Cargar archivo base
    if not os.path.exists(ARCHIVO_BASE_PATH):
        st.error(f"No se encontró el archivo base en: {ARCHIVO_BASE_PATH}")
        return None

    wb = load_workbook(ARCHIVO_BASE_PATH)

    if "Plantilla" not in wb.sheetnames or "Remoto" not in wb.sheetnames:
        st.error("El archivo base debe tener las hojas 'Plantilla' y 'Remoto'")
        return None

    ws_plantilla = wb["Plantilla"]
    ws_remoto = wb["Remoto"]

    # Limpiar Plantilla
    for row in ws_plantilla.iter_rows():
        for cell in row:
            cell.value = None

    # Escribir encabezados y datos
    for col_idx, col_name in enumerate(df_csv.columns, start=1):
        ws_plantilla.cell(row=1, column=col_idx, value=col_name)

    for row_idx, row in enumerate(df_csv.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws_plantilla.cell(row=row_idx, column=col_idx, value=value)

    # Limpiar columna B en Remoto
    for row in range(3, ws_remoto.max_row + 1):
        ws_remoto[f"B{row}"] = None

    # Escribir agentes en Remoto
    for i, nombre in enumerate(agentes_turno, start=3):
        ws_remoto[f"B{i}"] = nombre

    # Guardar en memoria
    output = BytesIO()
    fecha_actual = datetime.datetime.now().strftime("%d-%m-%Y")
    nombre_archivo = f"{fecha_actual}_{turno}.xlsx"
    wb.save(output)
    output.seek(0)
    return output, nombre_archivo

# =====================
# Interfaz Streamlit
# =====================
st.set_page_config(page_title="Stats Generator por Turno", page_icon="📊", layout="centered")
st.title("📊 Stats Generator por Turno")

horario = cargar_horario()

# Subir CSV LiveAgent
st.subheader("📂 Sube el CSV exportado de LiveAgent")
csv_file = st.file_uploader("Selecciona el archivo CSV", type=["csv"], key="liveagent")

col1, col2 = st.columns(2)
with col1:
    dia = st.selectbox("Día de la semana", sorted(horario["Día"].unique()))
with col2:
    turnos = horario[horario["Día"] == dia]["Turno"].unique()
    turno = st.selectbox("Turno", sorted(turnos))

# Preview de agentes si se subió CSV
if csv_file:
    try:
        df_csv = pd.read_csv(csv_file, encoding="utf-8-sig", sep=None, engine="python")
        # Llamar a generar_excel solo para preview
        generar_excel(horario, df_csv, dia, turno)
    except Exception as e:
        st.error(f"Error al leer CSV: {e}")

# Botón para generar Excel
if st.button("⚡ Generar Excel") and csv_file:
    try:
        df_csv = pd.read_csv(csv_file, encoding="utf-8-sig", sep=None, engine="python")
        excel_output, nombre_archivo = generar_excel(horario, df_csv, dia, turno)
        if excel_output:
            st.success("✅ Archivo generado con éxito")
            st.download_button(
                label="📥 Descargar Excel",
                data=excel_output,
                file_name=nombre_archivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    except Exception as e:
        st.error(f"Error al procesar: {e}")

# =====================
# Admin: Actualizar Horario (al fondo, desplegable)
# =====================
with st.expander("⚠️ Admin: Actualizar Horario (SOLO TOCAR DANIEL 👀)"):
    st.markdown("⚠️ Si lo toca alguien más, puedes cagar todo el bot")
    nuevo_horario = st.file_uploader("Subir nuevo horario.xlsx", type=["xlsx"], key="admin")
    
    if nuevo_horario:
        st.warning("⚠️ Has subido un archivo nuevo. Debes confirmar para reemplazar el horario actual.")
        confirmar = st.button("✅ Confirmar actualización del horario")
        if confirmar:
            # Crear backup del horario actual
            if os.path.exists(HORARIO_PATH):
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_path = os.path.join(BACKUP_FOLDER, f"horario_backup_{timestamp}.xlsx")
                shutil.copy2(HORARIO_PATH, backup_path)
            
            # Guardar nuevo horario
            with open(HORARIO_PATH, "wb") as f:
                f.write(nuevo_horario.getbuffer())
            st.success("✅ Horario actualizado correctamente y backup creado. Recarga la app.")
            st.stop()
